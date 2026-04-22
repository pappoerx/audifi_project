import csv
import io
import re
from datetime import datetime, timedelta

from django.contrib.auth import authenticate
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods, require_POST

from activity.models import ActivityEvent, ActivityType
from bookings.models import Course, TimeSlot, TutorialBooking, TutorialBookingStatus
from bookings.services import check_hall_availability, official_conflict, tutorial_conflict
from halls.models import Hall
from reports.models import IssueReport
from timetable.models import OfficialClass, WeekDay
from users.models import User, UserRole, UserToken

from .auth import parse_json_body, require_bearer_auth


def _serialize_user(user):
    return {
        "id": user.id,
        "institutional_id": user.institutional_id,
        "role": user.role,
        "display_name": user.display_name or user.full_name or user.username,
        "department": user.department,
        "program": user.program,
        "preferences": user.preferences or {},
    }


@require_POST
def auth_login(request):
    payload = parse_json_body(request)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON body."}, status=400)
    institutional_id = str(payload.get("institutional_id", "")).strip()
    password = payload.get("password", "")
    if not institutional_id or not password:
        return JsonResponse({"detail": "institutional_id and password are required."}, status=400)

    user = authenticate(request, username=institutional_id, password=password)
    if user is None:
        return JsonResponse({"detail": "Invalid credentials."}, status=401)

    token = UserToken.issue_for_user(user)
    return JsonResponse({"access_token": token.token, "user": _serialize_user(user)})


@require_http_methods(["GET", "PATCH"])
@require_bearer_auth()
def auth_me(request):
    user = request.auth_user
    if request.method == "GET":
        return JsonResponse(_serialize_user(user))

    payload = parse_json_body(request)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON body."}, status=400)
    if "display_name" in payload:
        user.display_name = payload.get("display_name") or ""
    if "department" in payload:
        user.department = payload.get("department")
    if "program" in payload:
        user.program = payload.get("program")
    if "preferences" in payload and isinstance(payload["preferences"], dict):
        user.preferences = payload["preferences"]
    user.save(update_fields=["display_name", "department", "program", "preferences"])
    return JsonResponse(_serialize_user(user))


@require_GET
@require_bearer_auth()
def halls_list(request):
    halls = Hall.objects.filter(is_active=True).order_by("name")
    only_available = request.GET.get("available_now") == "true"
    now = timezone.localtime()
    out = []
    instant_end = (now + timedelta(minutes=1)).time()
    for hall in halls:
        current_official = official_conflict(hall, now.date(), now.time(), instant_end)
        current_tutorial = tutorial_conflict(hall, now.date(), now.time(), instant_end)
        if current_official:
            status = "Occupied"
            event = current_official.course_name
            live = True
        elif current_tutorial:
            status = "Occupied"
            event = current_tutorial.course.title
            live = True
        else:
            future_tutorial = (
                TutorialBooking.objects.filter(
                    hall=hall,
                    booking_date=now.date(),
                    time_slot__start_time__gt=now.time(),
                    status__in=[TutorialBookingStatus.BOOKED, TutorialBookingStatus.IN_SESSION],
                )
                .select_related("course", "time_slot")
                .order_by("time_slot__start_time")
                .first()
            )
            if future_tutorial:
                status = "Booked - Pending"
                event = future_tutorial.course.title
                live = False
            else:
                status = "Available"
                event = "No active session"
                live = False

        if only_available and status != "Available":
            continue

        out.append(
            {
                "id": str(hall.id),
                "name": hall.name,
                "capacity": hall.capacity,
                "campus_zone": hall.campus_zone,
                "status": status,
                "current_or_next_event": event,
                "has_wifi": hall.has_wifi,
                "has_projector": hall.has_projector,
                "has_ac": hall.has_ac,
                "live": live,
            }
        )
    return JsonResponse(out, safe=False)


@require_GET
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def courses_list(request):
    rows = [{"id": str(c.id), "title": c.title, "code": c.code} for c in Course.objects.all()]
    return JsonResponse(rows, safe=False)


@require_GET
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def time_slots_list(request):
    rows = [
        {
            "id": str(s.id),
            "label": s.label,
            "start_time": s.start_time.strftime("%H:%M"),
            "end_time": s.end_time.strftime("%H:%M"),
        }
        for s in TimeSlot.objects.all()
    ]
    return JsonResponse(rows, safe=False)


@require_GET
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def halls_availability_for_slot(request):
    date_str = (request.GET.get("booking_date") or "").strip()
    time_slot_id = (request.GET.get("time_slot_id") or "").strip()
    start_time_str = (request.GET.get("start_time") or "").strip()
    end_time_str = (request.GET.get("end_time") or "").strip()
    if not date_str:
        return JsonResponse({"detail": "booking_date is required."}, status=400)
    if not time_slot_id and (not start_time_str or not end_time_str):
        return JsonResponse({"detail": "Provide either time_slot_id or start_time/end_time."}, status=400)
    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"detail": "Invalid booking_date format. Use YYYY-MM-DD."}, status=400)
    if time_slot_id:
        slot = TimeSlot.objects.filter(id=time_slot_id).first()
        if not slot:
            return JsonResponse({"detail": "Invalid time_slot_id."}, status=400)
        start_time = slot.start_time
        end_time = slot.end_time
        slot_label = slot.label
        slot_id_out = str(slot.id)
    else:
        try:
            start_time = datetime.strptime(start_time_str, "%H:%M").time()
            end_time = datetime.strptime(end_time_str, "%H:%M").time()
        except ValueError:
            return JsonResponse({"detail": "Invalid start_time/end_time format. Use HH:MM."}, status=400)
        if start_time >= end_time:
            return JsonResponse({"detail": "end_time must be later than start_time."}, status=400)
        slot_label = f"{start_time.strftime('%I:%M %p')} - {end_time.strftime('%I:%M %p')}"
        slot_id_out = ""
    if target_date < timezone.localdate():
        return JsonResponse({"detail": "Booking date has already passed."}, status=400)
    if target_date == timezone.localdate():
        now_time = timezone.localtime().time()
        if start_time <= now_time:
            return JsonResponse({"detail": "You cannot select a time that has already started."}, status=400)

    halls = Hall.objects.filter(is_active=True).order_by("name")
    available = []
    unavailable = []
    for hall in halls:
        status = check_hall_availability(hall, target_date, start_time, end_time)
        hall_name_upper = (hall.name or "").upper()
        if "LAB" in hall_name_upper:
            hall_type = "Lab"
        elif (
            "BLK" in hall_name_upper
            or "BLOCK" in hall_name_upper
            or "CLASS" in hall_name_upper
            or "AY" in hall_name_upper
            or "FF" in hall_name_upper
            or "SF" in hall_name_upper
            or "TF" in hall_name_upper
        ):
            hall_type = "Classroom"
        else:
            hall_type = "Lecture Hall"
        payload = {
            "id": str(hall.id),
            "name": hall.name,
            "capacity": hall.capacity,
            "has_ac": hall.has_ac,
            "has_projector": hall.has_projector,
            "has_display": hall.has_projector,
            "has_monitor": hall.has_projector,
            "has_audio_system": hall.has_projector,
            "has_microphone": hall.has_projector,
            "has_recording_capability": False,
            "type": hall_type,
        }
        if status["available"]:
            available.append(payload)
        else:
            unavailable.append({**payload, "reason": status["reason"]})

    return JsonResponse(
        {
            "booking_date": target_date.isoformat(),
            "time_slot_id": slot_id_out,
            "time_slot_label": slot_label,
            "available_halls": available,
            "unavailable_halls": unavailable,
        }
    )


@require_GET
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def fixed_timetable_list(request):
    rows = (
        OfficialClass.objects.filter(is_active=True)
        .select_related("hall", "lecturer")
        .order_by("day_of_week", "start_time", "hall__name")
    )
    data = [
        {
            "id": str(row.id),
            "course_name": row.course_name,
            "student_group": row.student_group,
            "hall_name": row.hall.name,
            "lecturer_name": row.lecturer_name or row.lecturer.display_name or row.lecturer.full_name or row.lecturer.username,
            "day_of_week": row.day_of_week,
            "start_time": row.start_time.strftime("%H:%M"),
            "end_time": row.end_time.strftime("%H:%M"),
            "semester": row.semester,
        }
        for row in rows
    ]
    return JsonResponse(data, safe=False)


@require_POST
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def fixed_timetable_upload(request):
    upload = request.FILES.get("file")
    if not upload:
        return JsonResponse({"detail": "Attach a CSV or XLSX file in the 'file' field."}, status=400)
    file_name = upload.name.lower()
    is_csv = file_name.endswith(".csv")
    is_xlsx = file_name.endswith(".xlsx")
    if not is_csv and not is_xlsx:
        return JsonResponse({"detail": "Only CSV and XLSX files are supported for web upload."}, status=400)

    semester_default = (request.POST.get("semester") or "").strip() or "Current"
    truncate = str(request.POST.get("truncate", "")).strip().lower() in {"1", "true", "yes", "on"}
    preview_only = str(request.POST.get("preview", "")).strip().lower() in {"1", "true", "yes", "on"}

    required = {"course_name", "hall_name", "day_of_week", "start_time", "end_time"}
    rows_data = []
    headers = []
    if is_csv:
        try:
            text_stream = io.TextIOWrapper(upload.file, encoding="utf-8-sig")
            reader = csv.DictReader(text_stream)
            headers = list(reader.fieldnames or [])
            rows_data = list(reader)
        except Exception:
            return JsonResponse({"detail": "Could not read uploaded CSV file."}, status=400)
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            return JsonResponse({"detail": "XLSX support requires openpyxl. Install dependencies and retry."}, status=500)
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
            sheet = wb.active
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if not first:
                return JsonResponse({"detail": "Uploaded XLSX is empty."}, status=400)
            headers = [str(x).strip() if x is not None else "" for x in first]
            for row_vals in iterator:
                row_map = {}
                has_value = False
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    val = row_vals[i] if i < len(row_vals) else None
                    if val is not None and str(val).strip() != "":
                        has_value = True
                    row_map[h] = "" if val is None else str(val).strip()
                if has_value:
                    rows_data.append(row_map)
        except Exception:
            return JsonResponse({"detail": "Could not read uploaded XLSX file."}, status=400)
        normalized_headers = {h.strip().lower() for h in headers if h}
        missing = required - normalized_headers
        if missing:
            # Fallback parser for grid-style timetable sheets (e.g., KSB format).
            rows_data = []
            day_map_short = {
                "mo": "monday",
                "tu": "tuesday",
                "we": "wednesday",
                "th": "thursday",
                "fr": "friday",
                "sa": "saturday",
                "su": "sunday",
            }
            day_full = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}
            time_re = re.compile(r"([01]?\d|2[0-3]):([0-5]\d)\s*[-–]\s*([01]?\d|2[0-3]):([0-5]\d)")
            course_re = re.compile(r"\b([A-Z]{2,6}\s?\d{3}[A-Z]?)\b")
            hall_re = re.compile(
                r"\b("
                r"SMA|VIRTUAL|BLK\s+[A-Z](?:\s+BASEMENT)?|BLOCK\s+[A-Z](?:\s*-\s*(?:MAIN|BASEMENT))?|"
                r"AY\s*(?:FF|SF|TF)\s*\d+|(?:FF|SF|TF)\s*\d+"
                r")\b",
                re.I,
            )
            student_group_re = re.compile(
                r"\b(BBA\s+[A-Z&]+(?:\s+[A-Z&]+)?\s*\d|BSC\s+[A-Z]+\s*\d|HTM\s*\d)\b",
                re.I,
            )

            def norm_time(hh, mm):
                return f"{int(hh):02d}:{mm}"

            def extract_groups(text):
                out = []
                for token in student_group_re.findall(text or ""):
                    norm = re.sub(r"\s+", " ", str(token).strip().upper())
                    norm = norm.replace("B& F", "B&F")
                    if norm and norm not in out:
                        out.append(norm)
                return out

            def extract_lecturer_name(text, course_code):
                lines = [re.sub(r"\s+", " ", ln).strip() for ln in str(text or "").splitlines() if str(ln).strip()]
                if not lines:
                    return ""
                course_norm = re.sub(r"\s+", "", str(course_code or "")).upper()
                idx = -1
                for i, ln in enumerate(lines):
                    if course_norm and course_norm in re.sub(r"\s+", "", ln).upper():
                        idx = i
                        break
                if idx >= 0 and idx + 1 < len(lines):
                    candidate = lines[idx + 1]
                else:
                    candidate = lines[1] if len(lines) > 1 else ""
                # Avoid picking hall labels or group lists as lecturer names.
                upper = candidate.upper()
                if (
                    upper.startswith("BLK")
                    or upper.startswith("SMA")
                    or upper.startswith("VIRTUAL")
                    or "BBA " in upper
                    or "BSC " in upper
                ):
                    return ""
                return candidate

            try:
                upload.file.seek(0)
                wb = load_workbook(upload, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    slot_by_col = {}
                    for row_vals in ws.iter_rows(values_only=True):
                        cells = ["" if v is None else str(v).strip() for v in row_vals]
                        # Update active time slots when a "8:00 - 10:00" header row appears.
                        found_time = False
                        for col_idx, cell in enumerate(cells):
                            m = time_re.search(cell)
                            if not m:
                                continue
                            slot_by_col[col_idx] = (norm_time(m.group(1), m.group(2)), norm_time(m.group(3), m.group(4)))
                            found_time = True
                        if found_time:
                            continue

                        if not cells:
                            continue
                        day_raw = (cells[0] or "").strip().lower()
                        if day_raw in day_map_short:
                            day_name = day_map_short[day_raw]
                        elif day_raw in day_full:
                            day_name = day_raw
                        else:
                            continue
                        if not slot_by_col:
                            continue

                        for col_idx in range(1, len(cells)):
                            cell_text = (cells[col_idx] or "").strip()
                            if not cell_text or col_idx not in slot_by_col:
                                continue
                            course_m = course_re.search(cell_text)
                            if not course_m:
                                continue
                            course_code = course_m.group(1).replace(" ", "")
                            hall_m = hall_re.search(cell_text)
                            hall_name = hall_m.group(1).upper() if hall_m else "TBA Hall"
                            lecturer_name = extract_lecturer_name(cell_text, course_code)
                            groups = extract_groups(cell_text) or ["UNSPECIFIED"]
                            for grp in groups:
                                rows_data.append(
                                    {
                                        "course_name": course_code,
                                        "student_group": grp,
                                        "lecturer_name": lecturer_name,
                                        "hall_name": re.sub(r"\s+", " ", hall_name),
                                        "day_of_week": day_name,
                                        "start_time": slot_by_col[col_idx][0],
                                        "end_time": slot_by_col[col_idx][1],
                                        "semester": semester_default,
                                    }
                                )
            except Exception:
                return JsonResponse(
                    {
                        "detail": "Could not parse this XLSX layout automatically. Use a tabular file with columns: "
                        "course_name,hall_name,day_of_week,start_time,end_time"
                    },
                    status=400,
                )

            if not rows_data:
                return JsonResponse(
                    {
                        "detail": "Could not parse this XLSX layout automatically. Use a tabular file with columns: "
                        "course_name,hall_name,day_of_week,start_time,end_time"
                    },
                    status=400,
                )
    if is_csv:
        missing = required - set(headers)
        if missing:
            return JsonResponse({"detail": "File is missing required columns: " + ", ".join(sorted(missing))}, status=400)

    if preview_only:
        preview_rows = []
        for row in rows_data[:30]:
            preview_rows.append(
                {
                    "course_name": (row.get("course_name") or "").strip(),
                    "student_group": (row.get("student_group") or "").strip(),
                    "lecturer_name": (row.get("lecturer_name") or "").strip(),
                    "hall_name": (row.get("hall_name") or "").strip(),
                    "day_of_week": (row.get("day_of_week") or "").strip(),
                    "start_time": (row.get("start_time") or "").strip(),
                    "end_time": (row.get("end_time") or "").strip(),
                    "semester": (row.get("semester") or "").strip() or semester_default,
                }
            )
        return JsonResponse(
            {
                "ok": True,
                "preview": True,
                "parsed_total": len(rows_data),
                "preview_rows": preview_rows,
            }
        )

    if truncate:
        OfficialClass.objects.all().delete()

    day_map = {
        "monday": WeekDay.MONDAY,
        "tuesday": WeekDay.TUESDAY,
        "wednesday": WeekDay.WEDNESDAY,
        "thursday": WeekDay.THURSDAY,
        "friday": WeekDay.FRIDAY,
        "saturday": WeekDay.SATURDAY,
        "sunday": WeekDay.SUNDAY,
    }
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows_data, start=2):
        course_name = (row.get("course_name") or "").strip()
        student_group = (row.get("student_group") or "").strip()
        lecturer_name = (row.get("lecturer_name") or "").strip()
        hall_name = (row.get("hall_name") or "").strip()
        day_raw = (row.get("day_of_week") or "").strip().lower()
        start_raw = (row.get("start_time") or "").strip()
        end_raw = (row.get("end_time") or "").strip()
        semester = (row.get("semester") or "").strip() or semester_default
        lecturer_id = (row.get("lecturer_id") or "").strip()

        if not all([course_name, hall_name, day_raw, start_raw, end_raw]):
            skipped += 1
            errors.append(f"Row {idx}: missing required value(s)")
            continue
        if day_raw not in day_map:
            skipped += 1
            errors.append(f"Row {idx}: invalid day_of_week '{day_raw}'")
            continue
        try:
            start_time = datetime.strptime(start_raw, "%H:%M").time()
            end_time = datetime.strptime(end_raw, "%H:%M").time()
        except ValueError:
            skipped += 1
            errors.append(f"Row {idx}: start_time/end_time must be HH:MM (24-hour)")
            continue
        if start_time >= end_time:
            skipped += 1
            errors.append(f"Row {idx}: start_time must be before end_time")
            continue

        hall = Hall.objects.filter(name=hall_name).first()
        if not hall:
            hall = Hall.objects.create(
                name=hall_name,
                capacity=200,
                location="School of Business",
                campus_zone="School of Business",
                has_wifi=True,
                has_projector=True,
                has_ac=False,
                is_active=True,
            )

        lecturer = request.auth_user
        if lecturer_id:
            lecturer = User.objects.filter(institutional_id=lecturer_id).first() or request.auth_user

        _, was_created = OfficialClass.objects.update_or_create(
            course_name=course_name,
            student_group=student_group,
            hall=hall,
            day_of_week=day_map[day_raw],
            start_time=start_time,
            end_time=end_time,
            semester=semester,
            defaults={"lecturer": lecturer, "lecturer_name": lecturer_name, "is_active": True},
        )
        if was_created:
            created += 1
        else:
            updated += 1

    return JsonResponse(
        {
            "ok": True,
            "preview": False,
            "parsed_total": len(rows_data),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "sample_errors": errors[:10],
        }
    )


def _serialize_booking(booking: TutorialBooking):
    return {
        "id": str(booking.id),
        "hall_name": booking.hall.name,
        "course_title": booking.course.title,
        "booking_date": booking.booking_date.isoformat(),
        "time_slot_label": booking.time_slot.label,
        "status": booking.status,
    }


def _create_activity(event_type, user, booking: TutorialBooking, note=""):
    ActivityEvent.objects.create(
        type=event_type,
        lecturer_name=user.display_name or user.full_name or user.username,
        auditorium=booking.hall.name,
        course=booking.course.title,
        date=booking.booking_date.isoformat(),
        time=booking.time_slot.label,
        note=note,
    )


@require_http_methods(["GET", "POST"])
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def bookings_resource(request):
    user = request.auth_user
    if request.method == "GET":
        rows = (
            TutorialBooking.objects.filter(booked_by=user)
            .select_related("hall", "course", "time_slot")
            .order_by("-booking_date", "time_slot__start_time")
        )
        return JsonResponse([_serialize_booking(b) for b in rows], safe=False)

    payload = parse_json_body(request)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON body."}, status=400)
    for field in ["hall_id", "course_id", "booking_date"]:
        if not payload.get(field):
            return JsonResponse({"detail": f"Missing field: {field}"}, status=400)
    time_slot_id = str(payload.get("time_slot_id") or "").strip()
    start_time_str = str(payload.get("start_time") or "").strip()
    end_time_str = str(payload.get("end_time") or "").strip()
    if not time_slot_id and (not start_time_str or not end_time_str):
        return JsonResponse({"detail": "Provide either time_slot_id or start_time/end_time."}, status=400)

    hall = Hall.objects.filter(id=payload["hall_id"], is_active=True).first()
    course = Course.objects.filter(id=payload["course_id"]).first()
    if not hall or not course:
        return JsonResponse({"detail": "Invalid hall/course."}, status=400)
    slot = None
    if time_slot_id:
        slot = TimeSlot.objects.filter(id=time_slot_id).first()
        if not slot:
            return JsonResponse({"detail": "Invalid time_slot_id."}, status=400)
        start_time = slot.start_time
        end_time = slot.end_time
    else:
        try:
            start_time = datetime.strptime(start_time_str, "%H:%M").time()
            end_time = datetime.strptime(end_time_str, "%H:%M").time()
        except ValueError:
            return JsonResponse({"detail": "Invalid start_time/end_time format. Use HH:MM."}, status=400)
        if start_time >= end_time:
            return JsonResponse({"detail": "end_time must be later than start_time."}, status=400)
        slot = (
            TimeSlot.objects.filter(start_time=start_time, end_time=end_time).order_by("id").first()
            or TimeSlot.objects.create(
                label=f"{start_time.strftime('%I:%M %p')} - {end_time.strftime('%I:%M %p')}",
                start_time=start_time,
                end_time=end_time,
            )
        )
    try:
        day = datetime.strptime(payload["booking_date"], "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"detail": "Invalid booking_date format."}, status=400)
    if day < timezone.localdate():
        return JsonResponse({"detail": "Booking date has already passed."}, status=400)
    if day == timezone.localdate():
        now_time = timezone.localtime().time()
        if start_time <= now_time:
            return JsonResponse({"detail": "You cannot book a time that has already started."}, status=400)

    availability = check_hall_availability(hall, day, start_time, end_time)
    if not availability["available"]:
        return JsonResponse({"detail": availability["reason"]}, status=409)

    booking = TutorialBooking.objects.create(
        hall=hall,
        course=course,
        booking_date=day,
        time_slot=slot,
        booked_by=user,
    )
    _create_activity(ActivityType.BOOKED, user, booking)
    return JsonResponse(_serialize_booking(booking), status=201)


@require_POST
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def cancel_booking(request, booking_id):
    user = request.auth_user
    booking = TutorialBooking.objects.filter(id=booking_id, booked_by=user).select_related("hall", "course", "time_slot").first()
    if not booking:
        return JsonResponse({"detail": "Booking not found."}, status=404)
    booking.status = TutorialBookingStatus.CANCELLED
    booking.save(update_fields=["status"])
    _create_activity(ActivityType.BOOKING_CANCELLED, user, booking)
    return JsonResponse({"ok": True})


@require_POST
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def call_off_booking(request, booking_id):
    user = request.auth_user
    booking = TutorialBooking.objects.filter(id=booking_id, booked_by=user).select_related("hall", "course", "time_slot").first()
    if not booking:
        return JsonResponse({"detail": "Booking not found."}, status=404)
    booking.status = TutorialBookingStatus.CANCELLED
    booking.save(update_fields=["status"])
    _create_activity(ActivityType.CLASS_CALLED_OFF, user, booking, note="Class called off")
    return JsonResponse({"ok": True})


@require_GET
@require_bearer_auth()
def activity_feed(request):
    limit = int(request.GET.get("limit", "12"))
    rows = ActivityEvent.objects.all()[: max(1, min(limit, 100))]
    data = [
        {
            "id": x.id,
            "at": x.at.isoformat(),
            "type": x.type,
            "lecturer_name": x.lecturer_name,
            "auditorium": x.auditorium,
            "course": x.course,
            "date": x.date,
            "time": x.time,
            "note": x.note,
        }
        for x in rows
    ]
    return JsonResponse(data, safe=False)


@require_GET
@require_bearer_auth(roles=[UserRole.STAFF, UserRole.TA, UserRole.ADMIN])
def staff_analytics(request):
    user = request.auth_user
    now = timezone.localdate()
    me_bookings = TutorialBooking.objects.filter(booked_by=user)
    active = me_bookings.filter(status__in=[TutorialBookingStatus.BOOKED, TutorialBookingStatus.IN_SESSION], booking_date__gte=now).count()
    today_sessions = me_bookings.filter(booking_date=now, status__in=[TutorialBookingStatus.BOOKED, TutorialBookingStatus.IN_SESSION]).count()
    week_start = now - timedelta(days=now.weekday())
    week_end = week_start + timedelta(days=6)
    week_sessions = me_bookings.filter(booking_date__range=[week_start, week_end]).count()
    distinct_halls = me_bookings.values("hall_id").distinct().count()
    catalog_halls = Hall.objects.filter(is_active=True).count()
    hall_cov = round((distinct_halls / catalog_halls) * 100, 1) if catalog_halls else 0

    events_30 = ActivityEvent.objects.filter(at__gte=timezone.now() - timedelta(days=30))
    events_7 = ActivityEvent.objects.filter(at__gte=timezone.now() - timedelta(days=7))
    booked30 = events_30.filter(type=ActivityType.BOOKED).count()
    cancelled30 = events_30.filter(type=ActivityType.BOOKING_CANCELLED).count()
    calledoff30 = events_30.filter(type=ActivityType.CLASS_CALLED_OFF).count()
    denom = max(booked30, 1)
    release_rate = round(((cancelled30 + calledoff30) / denom) * 100, 1)

    return JsonResponse(
        {
            "today_str": now.isoformat(),
            "active_reservations": active,
            "today_sessions": today_sessions,
            "week_sessions": week_sessions,
            "distinct_halls": distinct_halls,
            "catalog_halls": catalog_halls,
            "hall_coverage_pct": hall_cov,
            "lecturer_events_7d": events_7.count(),
            "keypad_checkins_30d": events_30.filter(type=ActivityType.CHECKED_IN_KEYPAD).count(),
            "new_bookings_logged_30d": booked30,
            "released_cancelled_30d": cancelled30,
            "classes_called_off_30d": calledoff30,
            "release_rate_display": f"{release_rate}%",
            "release_rate_hint": "Lower is better. Includes cancellations + call-offs against new bookings.",
            "release_rate_warn": release_rate > 25,
        }
    )


@require_POST
@require_bearer_auth()
def create_issue_report(request):
    user = request.auth_user
    payload = parse_json_body(request)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON body."}, status=400)
    if not payload.get("category") or not payload.get("description"):
        return JsonResponse({"detail": "category and description are required."}, status=400)
    hall = Hall.objects.filter(is_active=True).first()
    issue_text = f"[{payload.get('category')}] {payload.get('description')}"
    if payload.get("location"):
        issue_text = f"{issue_text} @ {payload.get('location')}"
    report = IssueReport.objects.create(
        student=user,
        hall=hall,
        issue=issue_text,
    )
    return JsonResponse({"id": report.id, "status": report.status}, status=201)
